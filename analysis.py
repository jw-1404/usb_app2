import sys
import os
import re
import struct
import time
import resources_rc
import threading
import queue
from collections import deque
from datetime import datetime
from typing import List, Dict, Optional
import pandas as pd
import openpyxl
import et_xmlfile
from dataclasses import dataclass
import numpy as np

# 导入 PyQt5 相关模块
from PyQt5.QtWidgets import QMainWindow, QApplication, QMessageBox, QFileDialog, QLabel
from PyQt5.QtCore import Qt, pyqtSignal, QThread, QObject, QTimer
from PyQt5.QtGui import QTextCursor, QTextCharFormat, QColor
from PyQt5.QtWidgets import QInputDialog, QProgressDialog
# 导入 USB 相关库
import usb.core
import usb.backend.libusb1
from ui_usb import Ui_MainWindow
# 设置全局 libusb 后端
backend = usb.backend.libusb1.get_backend(find_library=lambda x: "libusb-1.0.dll")




@dataclass
class Entry:
    trigger_id: int
    time_stamp: int
    sync_id: int
    sync_id_body: int
    fec_index: int
    ids: List[int]
    adcs: List[int]
# ID编码/解码函数
def encode_id(fec_index: int, chip_id: int, channel_id: int) -> int:
    # 假设3位fec,5位chip,8位channel（根据s_gid_idx<3,5,8>）
    return (fec_index << 13) | (chip_id << 8) | channel_id

def decode_id(gid: int) -> tuple:
    channel_id = gid & 0xFF  # 8位
    chip_id = (gid >> 8) & 0x1F  # 5位
    fec_index = (gid >> 13) & 0x07  # 3位
    return fec_index, chip_id, channel_id

# 二进制转字符串（类似to_binary_str）
def to_binary_str(value: int) -> str:

    return '0b' + format(value, '08b')
class AnalysisThread(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, dat_path, output_path, layer_sigma_map, chips_per_board=6):
        super().__init__()
        self.dat_path = dat_path
        self.output_path = output_path
        self.layer_sigma_map = layer_sigma_map  # 改为接收字典 {board_id: sigma_value}
        self.chips_per_board = chips_per_board  # 接收并保存芯片数量

    def run(self):
        try:
            def progress_cb(p):
                self.progress.emit(int(p))
                
            print(f"开始解析文件: {self.dat_path}")
            
            entries = self.unpack_dat(self.dat_path, progress_cb)
            
            if not entries:
                self.error.emit("无法从文件中提取有效数据")
                return
                
            print(f"解析完成，找到 {len(entries)} 个条目")
            
            ms_map = self.ana_baseline(entries)
            print(f"基线分析完成，ms_map 包含 {len(ms_map)} 个通道")
            
            if not ms_map:
                self.error.emit("基线分析失败：未计算出有效阈值")
                return
                
            self.generate_config(self.output_path, ms_map, self.layer_sigma_map)
            self.finished.emit(f"阈值文件生成完成，共处理 {len(entries)} 个数据帧")
            
        except Exception as e:
            error_msg = f"分析过程中发生错误: {str(e)}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            self.error.emit(error_msg)

    def unpack_dat(self, file_path: str, progress_callback=None) -> List[Entry]:
        """解析大端序二进制数据文件"""
        try:
            with open(file_path, 'rb') as f:
                data = f.read()
        except Exception as e:
            print(f"读取文件失败: {e}")
            return []
        
        entries = []
        pos = 0
        total_size = len(data)
        
        print(f"开始解析大端序文件，总大小: {total_size} 字节")
        
        HEAD_SIZE = 12  # start_tag(4) + frame_index(2) + reserved(1) + fec_index(1) + sync_id(4)
        BODY_START_TAG = 0xac0f
        BODY_END_TAG = 0xfffeaaaa
        CHIP_START_TAG = 0xfa
        CHIP_CHANNEL_NUMBER = 128
        CHIP_END0_MASK = 0x81000000
        CHIP_END1_MASK = 0x82000000
        CHIP_END2_MASK = 0x83000000
        TAIL_START_TAG = 0xffcc0000
        TAIL_SIZE = 12
        
        while pos + HEAD_SIZE < total_size:
            head_start_tag = struct.unpack_from('>I', data, pos)[0]
            if head_start_tag not in [0xffab530b, 0xffab530f]:
                pos += 1
                continue
            
            frame_index = struct.unpack_from('>H', data, pos + 4)[0]
            reserved = data[pos + 6]
            fec_index = data[pos + 7]
            sync_id = struct.unpack_from('>I', data, pos + 8)[0]
            
            if reserved != 0xd4 or fec_index >= 0x08:
                pos += 1
                continue
            
            pos += HEAD_SIZE
            
            if pos + 20 > total_size:
                break
            
            body_start_tag = struct.unpack_from('>H', data, pos)[0]
            if body_start_tag != BODY_START_TAG:
                pos += 1 - HEAD_SIZE
                continue
            
            frame_length = struct.unpack_from('>H', data, pos + 2)[0]
            trigger_id = struct.unpack_from('>I', data, pos + 4)[0]
            body_sync_id = struct.unpack_from('>I', data, pos + 8)[0]
            byte_length = struct.unpack_from('>I', data, pos + 12)[0]
            
            if byte_length / 4 + 5 != frame_length:
                pos += 1 - HEAD_SIZE
                continue
            
            pos += 16
            
            ids = []
            adcs = []
            
            chip_count = 0
            max_chips = 24
            
            while chip_count < max_chips and pos + 8 < total_size:
                chip_start_tag = data[pos]
                if chip_start_tag != CHIP_START_TAG:
                    break
                
                chip_id = data[pos + 1]
                ram_id = data[pos + 2]
                chip_word_number = data[pos + 3]
                reserved_chip = struct.unpack_from('>H', data, pos + 4)[0]
                temperature = struct.unpack_from('>H', data, pos + 6)[0]
                
                pos += 8
                
                for _ in range(CHIP_CHANNEL_NUMBER):
                    if pos + 4 > total_size:
                        break
                    value = struct.unpack_from('>I', data, pos)[0]
                    channel_id = (value >> 24) & 0xFF
                    adc = value & 0x00FFFFFF
                    if channel_id > 0x80:
                        break
                    gid = encode_id(fec_index, chip_id, channel_id)
                    ids.append(gid)
                    adcs.append(adc)
                    pos += 4
                
                if pos + 12 > total_size:
                    break
                end0 = struct.unpack_from('>I', data, pos)[0]
                end1 = struct.unpack_from('>I', data, pos + 4)[0]
                end2 = struct.unpack_from('>I', data, pos + 8)[0]
                if (end0 & 0xFF000000) != CHIP_END0_MASK or \
                   (end1 & 0xFF000000) != CHIP_END1_MASK or \
                   (end2 & 0xFF000000) != CHIP_END2_MASK:
                    break
                pos += 12
                chip_count += 1
            
            body_end_tag = struct.unpack_from('>I', data, pos)[0]
            if body_end_tag != BODY_END_TAG:
                pos += 1 - HEAD_SIZE - 16
                continue
            pos += 4
            
            while pos < total_size and data[pos] == 0x00:
                pos += 1
            
            if pos + TAIL_SIZE > total_size:
                break
            time_stamp = struct.unpack_from('>I', data, pos)[0]
            tail_start_tag = struct.unpack_from('>I', data, pos + 4)[0]
            crc32 = struct.unpack_from('>I', data, pos + 8)[0]
            
            if tail_start_tag != TAIL_START_TAG:
                pos += 1 - HEAD_SIZE - 16 - 4
                continue
            
            pos += TAIL_SIZE
            
            if adcs:
                entry = Entry(
                    trigger_id=trigger_id,
                    time_stamp=time_stamp,
                    sync_id=sync_id,
                    sync_id_body=body_sync_id,
                    fec_index=fec_index,
                    ids=ids,
                    adcs=adcs
                )
                entries.append(entry)
            
            if progress_callback:
                progress = (pos / total_size) * 100
                progress_callback(progress)
        
        return entries

    def ana_baseline(self, entries: List[Entry]) -> dict:
        """分析基线数据"""
        baseline_map = {}
        for entry in entries:
            for i, gid in enumerate(entry.ids):
                adc = entry.adcs[i]
                if gid not in baseline_map:
                    baseline_map[gid] = []
                baseline_map[gid].append(adc)
        
        ms_map = {}
        for gid, adcs_list in baseline_map.items():
            if len(adcs_list) > 0:
                arr = np.array(adcs_list)
                ms_map[gid] = (np.mean(arr), np.var(arr))
        
        return ms_map

    def generate_config(self, output_path: str, ms_map: dict, layer_sigma_map: dict):
        chips_per = self.chips_per_board
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            
            if not ms_map:
                print("警告: ms_map 为空，创建空文件")
                wb.save(output_path)
                return
            
            print(f"生成配置，共 {len(ms_map)} 个通道")
            
            def make_write_idx(gid):
                ids = decode_id(gid)  # (fec_index, chip_id, channel_id)
                row = ids[2]   # channel_id
                col = ids[0] * chips_per + ids[1] + 1
                if row < 1 or col < 1:
                    print(f"无效索引: gid={gid}, ids={ids}, row={row}, col={col}")
                return row, col
            
            def write_head_info(row_start, col, lid, range_str):
                lid_bit = 1 << lid
                lidx = f"B-{to_binary_str(lid_bit)}"
                ws.cell(row_start, col).value = lidx
                ws.cell(row_start + 1, col).value = "02-Reg02"
                for i in range(6):
                    ws.cell(row_start + 2 + i, col).value = "0x00ac0837"
                ws.cell(row_start + 8, col).value = "02-Reg03"
                for i in range(6):
                    ws.cell(row_start + 9 + i, col).value = "0x0080010f"
                ws.cell(row_start + 15, col).value = "04-Th_value"
                ws.cell(row_start + 16, col).value = "0x07"
                ws.cell(row_start + 17, col).value = "05-Th_enable"
                ws.cell(row_start + 18, col).value = "0x01"
                ws.cell(row_start + 19, col).value = "08-Filter"
                ws.cell(row_start + 20, col).value = "0x00"
                ws.cell(row_start + 21, col).value = "03-Th_config"
                start_row = row_start + 22
                for i in range(chips_per):
                    title = f"C-{to_binary_str(i)}"
                    ws.cell(start_row, col).value = title
                    ws.cell(start_row + 1, col).value = range_str % (i, i)
                    start_row += 2
            
            written_count = 0
            # 收集所有出现的 fec_index
            fec_indices_set = set()
            
            for gid in ms_map:
                fec_index, chip_id, channel_id = decode_id(gid)
                fec_indices_set.add(fec_index)
            
            fec_indices = sorted(list(fec_indices_set))
            print(f"检测到 FEC 索引: {fec_indices}")
            
            for gid in ms_map:
                mean, var = ms_map[gid]
                if var <= 0 or np.isnan(var) or np.isnan(mean):
                    print(f"跳过无效统计数据: gid={gid}, mean={mean}, var={var}")
                    continue
                
                # 根据板卡ID获取对应的Sigma值
                fec_index, chip_id, channel_id = decode_id(gid)
                board_id = 1 << fec_index  # 将fec_index转换为板卡ID
                
                # 获取该板卡对应的Sigma值，如果没有则使用默认值3.0
                sigmac = layer_sigma_map.get(board_id, 3.0)
                
                sigma = sigmac * np.sqrt(var)
                #threshold = int(np.floor(mean + sigma)) - 16384  # 减去十进制的16384（即0x4000）
                threshold = int(np.floor(16384-sigma))
                if threshold < 0:
                    threshold = 0
                
                row, col = make_write_idx(gid)
                if row < 1 or col < 1:
                    print(f"无效索引: gid={gid}, row={row}, col={col}")
                    continue
                ws.cell(row=row, column=col).value = f"0x{threshold:08x}"
                written_count += 1
            
            # 动态生成 write_head_info，基于实际的 fec_index
            range_strs = [
                "${0:A+%d,128:B+%d}",
                "${0:A+24+%d,128:B+24+%d}",
                "${0:A+48+%d,128:B+48+%d}",
                "${0:A+72+%d,128:B+72+%d}",
                "${0:A+96+%d,128:B+96+%d}",
                "${0:A+120+%d,128:B+120+%d}",
                "${0:A+144+%d,128:B+144+%d}",
                "${0:A+168+%d,128:B+168+%d}"
            ]
            
            for col_idx, fec_idx in enumerate(fec_indices):
                col = col_idx + 1
                range_str = range_strs[fec_idx] if fec_idx < len(range_strs) else range_strs[0]
                write_head_info(129, col, fec_idx, range_str)
            
             # 将“最后一层”之前的所有列中的空单元格填充为 0x00000000
            # 先计算已写入阈值的最大列（对应最后一层使用到的列）
            max_col = 0
            for gid in ms_map:
                fidx, chipid, _ = decode_id(gid)
                col_idx = fidx * chips_per + chipid + 1
                if col_idx > max_col:
                    max_col = col_idx

            # 如果找到了最大列，则把 1 .. max_col-1 的空单元格都填成 0x00000000
            if max_col > 1:
                for col in range(1, max_col):
                    for row in range(1, 129):  # 通道0-127对应行1-128
                        if ws.cell(row=row, column=col).value is None:
                            ws.cell(row=row, column=col).value = "0x00000000"
            print(f"成功写入 {written_count} 个阈值数据")
            wb.save(output_path)
            print(f"配置文件已保存: {output_path}")
            
        except Exception as e:
            print(f"生成配置文件失败: {e}")
            raise
